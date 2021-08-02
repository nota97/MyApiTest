# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import os
import testdata.ExcelConfig as ExcelConfig
import requests
import json
import openpyxl

#读取项目路径
path = os.path.split(os.path.realpath(__file__))[0]


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('PyCharm')
    print(os.path.join(path,'data'))
    print(ExcelConfig.ExcelName,ExcelConfig.ExcelSheet)
    # url="http://192.168.1.6:31046/hermes/device/loginTerminalInfo/{identifier}"
    # data="{'identifier':'0ecc5f6d41b2311fca1230b62ee78f0a_1','username':'xxxx'}"
    # print(type(data))
    # res = requests.get(url=url,
    #                    params=data).json()
    #
    # r = json.dumps(res, ensure_ascii=False, sort_keys=True, indent=2)
    # if (type(data) is not dict):
    #     print("11111")
    # else:
    #     print("22222222")
    xpath = os.path.join(path+'/testdata', ExcelConfig.ExcelName)
    file = openpyxl.load_workbook(xpath)
    sheet = file[ExcelConfig.ExcelSheet]
    rows = sheet.max_row
    col = sheet.max_column
    sheet.cell(5, 1).value="http://192.168.1.27:12753/hermes/device/getNetworkConfig/{identifier}"
    sheet.cell(5, 2).value="get"
    sheet.cell(5, 3).value=""
    sheet.cell(5, 4).value="{'identifier':'13111111111'}"
    sheet.cell(5, 5).value="E.NETWORK.CONFIG.DEPARTPATH.NOTEXIST"

    file.save(path+'/testdata/'+ExcelConfig.ExcelName)






# See PyCharm help at https://www.jetbrains.com/help/pycharm/
